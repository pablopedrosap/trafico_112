from __future__ import annotations

"""
Batch processor for Clínica CEMTRO Tráfico.
------------------------------------------------
1. Lee automáticamente los correos reenviados a *trafico@atencionpericial.com* con facturas (FAC) y evolutivos (EVO) en PDF/JPG/PNG.
2. Agrupa adjuntos por **N° de historia / DNI + fecha** → crea carpeta → guarda pareja FAC‑EVO.
3. Si el PDF contiene varias páginas o es imagen escaneada, usa **Gemini Flash Vision** para extraer texto (con reintentos) y verifica campos clave.
4. Rellena/actualiza un `trafico_master.xlsx` con: Paciente, Importe total, #Facturas, #Evolutivos, Incidencias.
5. Genera `incidencias.csv` con todo lo que falte o no cuadre.
6. Responde al usuario con un *.zip* que incluye carpetas por paciente, el Excel, el CSV y *pairing_prompt.txt*.

Uso ▶  `python trafico_processor.py --once`  
     ▶  `python trafico_processor.py --loop`  (cada 10 min)
"""

# --------------------------------------------------------------------------------------
# IMPORTS
# --------------------------------------------------------------------------------------
import argparse
import asyncio
import base64
import email
import imaplib
import json
import os
import re
import smtplib
import textwrap
import time
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Dict, List, Tuple

import fitz  # PyMuPDF
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook

# --------------------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------------------
MIME_MAP = {".pdf": "application/pdf", ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}

load_dotenv()
IMAP_HOST, IMAP_USER, IMAP_PASS = os.getenv("IMAP_HOST"), os.getenv("IMAP_USER"), os.getenv("IMAP_PASS")
SMTP_HOST, SMTP_USER, SMTP_PASS = os.getenv("SMTP_HOST"), os.getenv("SMTP_USER"), os.getenv("SMTP_PASS")
GEMINI_KEY = os.getenv("GEMINI_KEY")
BASE_DIR = Path(os.getenv("DATA_DIR", "data"))
BASE_DIR.mkdir(exist_ok=True)
MASTER_XLSX = BASE_DIR / "trafico_master.xlsx"
INCID_CSV = BASE_DIR / "incidencias.csv"
PROMPT_TXT = BASE_DIR / "pairing_prompt.txt"
MODEL_NAME_1 = "models/gemini-2.0-flash"
MODEL_NAME_2 = "models/gemini-2.0-flash"
MAX_OUTPUT_TOKENS = 20000
MAX_CONCURRENCY = 8

# --------------------------------------------------------------------------------------
# GEMINI CLIENT
# --------------------------------------------------------------------------------------
import google.generativeai as clientgem
clientgem.configure(api_key=GEMINI_KEY)


async def _retry(fn, *args, **kwargs):
    delays = (0, 60, 180)
    last_exc: Exception | None = None
    for i, d in enumerate(delays):
        if d:
            await asyncio.sleep(d)
        try:
            return await asyncio.to_thread(fn, *args, **kwargs)
        except Exception as e:  # pragma: no cover
            last_exc = e
            if i == len(delays) - 1:
                raise
            print(f"[Gemini] intento {i+1} falló: {e} — reintento en {delays[i+1]} s")
    raise last_exc  # type: ignore

# --------------------------------------------------------------------------------------
# 1️⃣ CLASIFICACIÓN INDIVIDUAL
# --------------------------------------------------------------------------------------

async def classify_file(path: Path) -> Dict[str, Any]:
    mime_type = MIME_MAP.get(path.suffix.lower(), "application/octet-stream")
    b64 = base64.b64encode(path.read_bytes()).decode()

    prompt = (
        "Eres un sistema pericial. Devuelve SOLO este JSON:\n"
        "{\"tipo_documento\":\"FACTURA|EVOLUTIVO|DESCONOCIDO\",\"nombre_paciente\":\"... (formato NOMBRE APELLIDO1 APELLIDO2)\",\"dni_paciente\":\"... (formato 01234567A)\",\"numero_episodio\":\"... (si lo hay)\",\"fecha_documento\":\"... (formato DD/MM/AAAA)\",\"importe\":\"... (solo si es factura, pon el total y el número sin euros para que se coja como float)\",\"detalles\":\"... (detallar todo lo relevante para emparejarlo con su factura si se trata de un evolutivo y viceversa. Todas las fechas relevantes deben incluirse)\",\"confianza\":0.9}\n\n"
    )
    model = clientgem.GenerativeModel(MODEL_NAME_1)
    resp = await _retry(
        model.generate_content,
        contents=[{"parts": [{"text": prompt}, {"inline_data": {"mime_type": mime_type, "data": b64}}]}],
        generation_config=clientgem.GenerationConfig(max_output_tokens=MAX_OUTPUT_TOKENS, temperature=0.1),
    )
    raw = resp.text
    j = raw[raw.find("{"): raw.rfind("}") + 1]
    return json.loads(j) | {"archivo_origen": path.name}


async def classify_batch(paths: List[Path]):
    sem = asyncio.Semaphore(MAX_CONCURRENCY)

    async def _w(p: Path):
        async with sem:
            print("🔎", p.name)
            return await classify_file(p)

    return [r for r in await asyncio.gather(*[asyncio.create_task(_w(p)) for p in paths]) if r]

# --------------------------------------------------------------------------------------
# 2️⃣ PAIRING PROMPT + LLM MATCH
# --------------------------------------------------------------------------------------

MAX_LINES, MAX_DETAILS = 400, 1200


def build_pairing_prompt(docs: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for d in docs[:MAX_LINES]:
        detalle = (d.get("detalles") or "").replace("\n", " ")
        if len(detalle) > MAX_DETAILS:
            detalle = detalle[:MAX_DETAILS] + "…"
        lines.append(
            f"- {d['tipo_documento'][:8]:<8} | {d['archivo_origen']} | {(d['nombre_paciente'] or 'SIN_NOMBRE').upper()} | {d['dni_paciente'] or 'SIN_DNI'} | {d['fecha_documento'] or 'SIN_FECHA'} | {detalle}"
        )

    prompt = f"""
Eres un ayudante experto en siniestros.
1. Agrupa documentos por paciente (nombres/DNI pueden tener variaciones pero siendo el mismo paciente, en el json final simplemente pon el nombre más común si hay variaciones).
2. Dentro de cada paciente crea **casos**. Un caso agrupa 1-N evolutivos y 1-N facturas que
   se refieren al mismo evento (no tiene porqué haber un número de episodio pero si lo hay facilita la identificación) (fecha ±3 días y descripción coincidente).
3. Devuelve SOLO JSON sin comentarios con la siguiente estructura exacta:
{{
  "pacientes": {{
    "PACIENTE": {{
      "casos": [{{"evolutivos": [xxx.pdf, ...], "facturas": [...]}}, ...],
      "evolutivos_sueltos": [...],
      "facturas_sueltas":  [xxx.pdf, ...],
      "justificación_de_faltantes": "Si hay documentos que no se han podido agrupar, indica por qué exactamente no lo has agrupado.",
    }}
  }}
}} 
  
### LISTA DE DOCUMENTOS
{chr(10).join(lines)}
"""
    return textwrap.dedent(prompt).strip()


def extract_json(raw: str):
    raw = raw.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```[^\n]*\n?|```$", "", raw, flags=re.M)
    m = re.search(r"\{.*\}", raw, re.S)
    if not m:
        raise ValueError("No JSON in LLM response")
    return json.loads(m.group())


async def pair_with_llm(prompt: str):
    model = clientgem.GenerativeModel(MODEL_NAME_2)
    resp = await _retry(model.generate_content, prompt)

    print("LLM response:", resp.text)  # Print first 1000 chars for debugging
    print("finished llm response")
    return extract_json(resp.text)["pacientes"]

# --------------------------------------------------------------------------------------
# 3️⃣ CONSTRUIR patient_groups A PARTIR DEL JSON
# --------------------------------------------------------------------------------------

async def build_patient_groups(paths: List[Path]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """Clasifica, llama al LLM para agrupar y devuelve un dict de grupos de paciente."""
    docs = await classify_batch(paths)
    prompt = build_pairing_prompt(docs)
    PROMPT_TXT.write_text(prompt, encoding="utf-8")

    try:
        pacientes_json = await pair_with_llm(prompt)
        print("✅ LLM pairing exitoso")
    except Exception as e:                       # si el LLM falla seguimos
        print("⚠️ LLM pairing failed:", e)
        pacientes_json = {}

    file_map = {d["archivo_origen"]: d for d in docs}
    referenced: set[str] = set()
    pgs: Dict[Tuple[str, str], Dict[str, Any]] = {}

    # --- recorre cada paciente del JSON ------------------------------------------------
    for nombre, pdata in pacientes_json.items():
        key = (nombre, "")
        g = pgs.setdefault(
            key,
            {
                "paciente": nombre,
                "dni": "",
                "casos": [],
                "facturas": [],
                "evolutivos": [],
                "documentos_desconocidos": [],
                "importe_total": 0.0,
            },
        )

        # 1) casos 1‑N evolutivos ↔ 1‑N facturas
        for caso in pdata.get("casos", []):
            c_obj = {"evolutivos": [], "facturas": []}

            for evo in caso.get("evolutivos", []):
                if evo in file_map:
                    doc = file_map[evo]
                    c_obj["evolutivos"].append(doc)
                    g["evolutivos"].append(doc)
                    referenced.add(evo)
                    if not g["dni"] and doc.get("dni_paciente"):
                        g["dni"] = doc["dni_paciente"]

            for fac in caso.get("facturas", []):
                if fac in file_map:
                    doc = file_map[fac]
                    c_obj["facturas"].append(doc)
                    g["facturas"].append(doc)
                    referenced.add(fac)
                    if not g["dni"] and doc.get("dni_paciente"):
                        g["dni"] = doc["dni_paciente"]
                    try:
                        g["importe_total"] += float(doc.get("importe") or 0)
                    except (ValueError, TypeError):
                        pass

            g["casos"].append(c_obj)

        g["from_llm"] = True

        # 2) sueltos (no asignados a ningún caso)
        for evo in pdata.get("evolutivos_sueltos", []):
            if evo in file_map:
                doc = file_map[evo]
                g["evolutivos"].append(doc)
                referenced.add(evo)
        for fac in pdata.get("facturas_sueltas", []):
            if fac in file_map:
                doc = file_map[fac]
                g["facturas"].append(doc)
                referenced.add(fac)
                try:
                    g["importe_total"] += float(doc.get("importe") or 0)
                except (ValueError, TypeError):
                    pass

    # 3) archivos que el LLM no mencionó → grupo SIN_PACIENTE
    for d in docs:
        if d["archivo_origen"] in referenced:
            continue

        # crea/recupera el grupo residual una sola vez
        residual = pgs.setdefault(
            ("OTRO_TIPO_DE_DOC_O_SIN_PACIENTE", ""),
            {
                "paciente": "OTRO_TIPO_DE_DOC_O_SIN_PACIENTE",
                "dni": "",
                "casos": [],
                "facturas": [],
                "evolutivos": [],
                "documentos_desconocidos": [],
                "importe_total": 0.0,
            },
        )
        residual["documentos_desconocidos"].append(d)
        residual["from_llm"] = False
        

    return pgs

# --------------------------------------------------------------------------------------
# 4️⃣ EXCEL, CSV, INCIDENCIAS
# --------------------------------------------------------------------------------------

from datetime import datetime

DATE_FMTS = ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y")

def _parse_date(s: str) -> datetime | None:
    for f in DATE_FMTS:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    return None

def _date_of_case(caso: dict[str, list[dict]]) -> str:
    dts = [
        _parse_date(d.get("fecha_documento"))
        for d in caso["evolutivos"] + caso["facturas"]
        if d.get("fecha_documento")
    ]
    dts = [dt for dt in dts if dt]
    if not dts:
        return "SIN_FECHA"
    # antes: "%d-%m-%Y"
    return min(dts).strftime("%Y-%m-%d")


def incidencias_from_groups(pgs: dict) -> list[str]:
    out: list[str] = []

    for g in pgs.values():
        if not g.get("from_llm"):
            continue

        for caso in g.get("casos", []):
            if not caso["facturas"]:      # evolutivos sin pareja
                for evo in caso["evolutivos"]:
                    out.append(f"{g['paciente']}: {evo['archivo_origen']} SIN_FACTURA")
            if not caso["evolutivos"]:    # facturas sin pareja
                for fac in caso["facturas"]:
                    out.append(f"{g['paciente']}: {fac['archivo_origen']} SIN_EVOLUTIVO")

        # documentos desconocidos
        for doc in g.get("documentos_desconocidos", []):
            out.append(f"{g['paciente']}: {doc['archivo_origen']} DESCONOCIDO")

    return sorted(set(out))


def update_excel(pgs: dict, incidencias: list[str]) -> None:
    rows = []
    for g in pgs.values():
        if not g.get("from_llm"):          # ← salta los grupos residuales
            continue

        inc = "; ".join(sorted({i for i in incidencias if i.startswith(g["paciente"] + ":")}) or ["OK"])

        rows.append([
            g["paciente"],                           # nombre canónico tal cual
            g["dni"],                                # el primero que hayamos captado
            len(g["facturas"]),
            g["importe_total"],
            len(g["evolutivos"]),
            inc,
        ])

    df = pd.DataFrame(rows, columns=[
        "Paciente", "DNI", "Num_Facturas",
        "Importe_Total", "Num_Evolutivos", "Incidencias"
    ])

    # Sobrescribe siempre el fichero (ya no se concatenan duplicados)
    wb = Workbook()
    ws = wb.active; ws.title = "Datos"
    ws.append(df.columns.tolist())
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Prompt como antes
    ws_prompt = wb.create_sheet("Prompt")
    for i, line in enumerate(PROMPT_TXT.read_text("utf-8").splitlines(), 1):
        ws_prompt.cell(i, 1, line)

    wb.save(MASTER_XLSX)



def save_csv(inc: List[str]):
    INCID_CSV.write_text("Incidencia\n" + "\n".join(inc or ["No hay incidencias"]), encoding="utf-8")

# --------------------------------------------------------------------------------------
# 5️⃣ ZIP & EMAIL
# --------------------------------------------------------------------------------------

def gen_filename(d: Dict[str, Any]):
    safe = re.sub(r"[^\w\s-]", "", d.get("nombre_paciente") or "SIN_NOMBRE")
    safe = re.sub(r"\s+", "_", safe)[:20]
    tipo = "FAC" if d["tipo_documento"] == "FACTURA" else "EVO" if d["tipo_documento"] == "EVOLUTIVO" else "DESC"
    fecha = (d.get("fecha_documento") or "SIN_FECHA").replace("-", "").replace("/", "")
    return f"{safe}_{tipo}_{fecha}{Path(d['archivo_origen']).suffix}"

import math, itertools, zipfile

MAX_ZIP_BYTES = 20 * 1024 * 1024      # 20 MB = 20 × 2^20 bytes

def _write_zip(zip_path: Path, entries: list[tuple[Path, str]]) -> None:
    """Escribe un zip DEFLATE nivel 9 con las (src, arcname) dadas."""
    with zipfile.ZipFile(zip_path, "w",
                         compression=zipfile.ZIP_DEFLATED,
                         compresslevel=9) as zf:
        for src, arc in entries:
            zf.write(src, arcname=arc)
        zf.write(MASTER_XLSX, MASTER_XLSX.name)
        # zf.write(INCID_CSV,  INCID_CSV.name)


def build_zips(pgs) -> list[Path]:
    """Crea zips ≤20 MB.  
       – Casos sin factura → carpeta INCIDENCIAS  
       – Docs huérfanos → carpeta INCIDENCIAS raiz
    """
    entries: list[tuple[Path, str]] = []

    for g in pgs.values():
        patient_folder = re.sub(r"\s+", "_", g["paciente"])[:30]

        # 1) casos detectados por el LLM
        for caso in g.get("casos", []):
            if caso["facturas"]:                       # ⚑ solo casos con FAC
                date_folder = _date_of_case(caso)
                for doc in itertools.chain(caso["facturas"], caso["evolutivos"]):
                    src = BASE_DIR / doc["archivo_origen"]
                    if src.exists():
                        arc = f"{patient_folder}/{date_folder}/{gen_filename(doc)}"
                        entries.append((src, arc))
            else:                                      # ⮡ sin FAC → incidencias
                for doc in caso["evolutivos"]:
                    src = BASE_DIR / doc["archivo_origen"]
                    if src.exists():
                        arc = f"{patient_folder}/INCIDENCIAS/{gen_filename(doc)}"
                        entries.append((src, arc))

        # 2) docs desconocidos del grupo
        for doc in g.get("documentos_desconocidos", []):
            src = BASE_DIR / doc["archivo_origen"]
            if src.exists():
                arc = f"{patient_folder}/INCIDENCIAS/{gen_filename(doc)}"
                entries.append((src, arc))

    # 3) docs sin paciente → carpeta INCIDENCIAS a nivel raíz
    for g in pgs.values():
        if g["paciente"] == "OTRO_TIPO_DE_DOC_O_SIN_PACIENTE":
            for doc in itertools.chain(
                g.get("evolutivos", []),
                g.get("facturas", []),
                g.get("documentos_desconocidos", []),
            ):
                src = BASE_DIR / doc["archivo_origen"]
                if src.exists():
                    arc = f"INCIDENCIAS/{gen_filename(doc)}"
                    entries.append((src, arc))

    # 2) Divide en “baldes” de tamaño ≤20 MB (suma de tamaños sin comprimir ≈ buena heurística)

    entries.sort(key=lambda t: t[1])   # t = (src_path, arcname)

    batches: list[list[tuple[Path, str]]] = []
    current, cur_bytes = [], 0

    for src, arc in entries:
        size = src.stat().st_size
        if cur_bytes + size > MAX_ZIP_BYTES and current:
            batches.append(current)
            current, cur_bytes = [], 0
        current.append((src, arc))
        cur_bytes += size
    if current:
        batches.append(current)

    # 3) Escribe cada zip y, si algún zip sigue >20 MB tras DEFLATE, re‑parte en zips de un solo fichero
    out_files: list[Path] = []
    for i, batch in enumerate(batches, 1):
        zpath = BASE_DIR / f"lote_{int(time.time())}_{i}.zip"
        _write_zip(zpath, batch)
        if zpath.stat().st_size > MAX_ZIP_BYTES:
            # re‑crear con lotes unitarios (rara vez necesario)
            zpath.unlink()
            for j, ent in enumerate(batch, 1):
                zp = BASE_DIR / f"lote_{int(time.time())}_{i}_{j}.zip"
                _write_zip(zp, [ent])
                out_files.append(zp)
        else:
            out_files.append(zpath)

    return out_files


MAX_MAIL_BYTES = 20 * 1024 * 1024        # 20 MB

def send_reply_multi(to_addr: str, zip_paths: list[Path]) -> None:
    """Manda un correo por cada ZIP (todos <20 MB)."""
    for idx, zp in enumerate(zip_paths, 1):
        size = zp.stat().st_size
        if size > MAX_MAIL_BYTES:
            raise RuntimeError(f"{zp.name} sigue pesando {size/1_048_576:.1f} MB")

        msg = EmailMessage()
        plural = "" if len(zip_paths) == 1 else f" (parte {idx}/{len(zip_paths)})"
        msg["Subject"] = f"Evolutivos procesados{plural}"
        msg["From"] = SMTP_USER
        msg["To"] = to_addr
        msg.set_content(
            "Adjunto ZIP con carpetas por paciente y el Excel.\n"
            "Si recibes varios correos, únelos; cada uno es independiente."
        )
        msg.add_attachment(
            zp.read_bytes(),
            maintype="application",
            subtype="zip",
            filename=zp.name
        )

        with smtplib.SMTP_SSL(SMTP_HOST) as s:
            s.login(SMTP_USER, SMTP_PASS)
            s.send_message(msg)
        print(f"📤 Enviado {zp.name} ({size/1_048_576:.1f} MB)")


# --------------------------------------------------------------------------------------
# 6️⃣ IMAP: DESCARGA DE EMAILS
# --------------------------------------------------------------------------------------

def connect_imap():
    imap = imaplib.IMAP4_SSL(IMAP_HOST)
    imap.login(IMAP_USER, IMAP_PASS)
    imap.select("INBOX")
    return imap


def fetch_unseen_with_attachments(imap):
    typ, data = imap.search(None, "UNSEEN")
    ids = data[0].split()
    for uid in ids:
        typ, msg_data = imap.fetch(uid, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        sender = email.utils.parseaddr(msg.get("From"))[1]
        atts: List[Path] = []
        for part in msg.walk():
            if part.get_content_maintype() == "multipart":
                continue
            fname = part.get_filename()
            if fname and fname.lower().endswith(tuple(MIME_MAP)):
                p = BASE_DIR / fname
                p.write_bytes(part.get_payload(decode=True))
                atts.append(p)
        if atts:
            yield sender, atts

# --------------------------------------------------------------------------------------
# 7️⃣ MAIN LOOP
# --------------------------------------------------------------------------------------

async def main_async(loop: bool):
    imap = connect_imap()
    try:
        while True:
            for sender, files in fetch_unseen_with_attachments(imap):
                print(f"📩 {sender} → {len(files)} adjuntos")
                pgs = await build_patient_groups(files)
                print("👥 grupos:", (pgs))

                inc = incidencias_from_groups(pgs)
                print("⚠️ incidencias:", (inc))
                update_excel(pgs, inc)
                # save_csv(inc)
                
                zip_paths = build_zips(pgs)
                send_reply_multi(sender, zip_paths)
                print("✅ Enviados:", ", ".join(z.name for z in zip_paths))

            if not loop:
                break
            time.sleep(600)
    finally:
        imap.logout()


def main(loop: bool):
    asyncio.run(main_async(loop))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--loop", action="store_true")
    main(loop=ap.parse_args().loop)





    # ✅ LLM pairing exitoso
# M. BLANCA ENTRENA PALOMERO {'pairs': [['Documento 4.pdf', 'Documento 33.pdf'], ['Documento 5.pdf', 'Documento 32.pdf'], ['Documento 5.pdf', 'Documento 33.pdf'], ['Documento 6.pdf', 'Documento 34.pdf'], ['Documento 7.pdf', 'Documento 35.pdf'], ['Documento 8.pdf', 'Documento 36.pdf'], ['Documento 9.pdf', 'Documento 37.pdf'], ['Documento 10.pdf', 'Documento 37.pdf'], ['Documento 11.pdf', 'Documento 38.pdf'], ['Documento 12.pdf', 'Documento 39.pdf'], ['Documento 13.pdf', 'Documento 39.pdf'], ['Documento 13.pdf', 'Documento 40.pdf'], ['Documento 14.pdf', 'Documento 41.pdf'], ['Documento 15.pdf', 'Documento 42.pdf'], ['Documento 16.pdf', 'Documento 43.pdf'], ['Documento 17.pdf', 'Documento 44.pdf'], ['Documento 18.pdf', 'Documento 44.pdf'], ['Documento 19.pdf', 'Documento 44.pdf'], ['Documento 20.pdf', 'Documento 45.pdf'], ['Documento 21.pdf', 'Documento 46.pdf'], ['Documento 22.pdf', 'Documento 47.pdf'], ['Documento 23.pdf', 'Documento 47.pdf'], ['Documento 24.pdf', 'Documento 48.pdf'], ['Documento 25.pdf', 'Documento 48.pdf'], ['Documento 26.pdf', 'Documento 49.pdf'], ['Documento 26.pdf', 'Documento 50.pdf'], ['Documento 27.pdf', 'Documento 49.pdf'], ['Documento 27.pdf', 'Documento 50.pdf'], ['Documento 28.pdf', 'Documento 49.pdf'], ['Documento 28.pdf', 'Documento 50.pdf'], ['Documento 29.pdf', 'Documento 49.pdf'], ['Documento 29.pdf', 'Documento 50.pdf'], ['Documento 30.pdf', 'Documento 51.pdf'], ['Documento 30.pdf', 'Documento 52.pdf'], ['Documento 30.pdf', 'Documento 53.pdf'], ['Documento 30.pdf', 'Documento 54.pdf'], ['Documento 31.pdf', 'Documento 51.pdf'], ['Documento 31.pdf', 'Documento 52.pdf'], ['Documento 31.pdf', 'Documento 53.pdf'], ['Documento 31.pdf', 'Documento 54.pdf']], 'unpaired_evo': ['Documento 3 BIS.pdf'], 'unpaired_fac': []}